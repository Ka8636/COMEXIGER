from django.db import transaction
from django.db import IntegrityError
from django.db.models import Sum
from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import datetime
from django.utils import timezone
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer

from rest_framework import viewsets, status
from rest_framework.decorators import api_view, action
from rest_framework.response import Response

from .serializers import DisponibilidadSerializer
from .models import Disponibilidad, QRDisponibilidadUsado,Variedad

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated


from .models import Variedad
from .serializers import VariedadSerializer
from Aplicaciones.Usuario.jwt_decorators import jwt_required
from django.db.models.deletion import ProtectedError


from Aplicaciones.Usuario.web_decorators import web_admin_required
from Aplicaciones.Usuario.models import Mesa, Usuario


def _to_positive_int(value):
    try:
        n = int(str(value).strip())
        return n if n > 0 else None
    except (TypeError, ValueError):
        return None


def _resolver_mesa_para_creacion(request, variedad, medida, mesa_raw):
    mesa = _to_positive_int(mesa_raw)
    if mesa:
        return mesa

    previo = (Disponibilidad.objects
              .filter(variedad=variedad, medida=medida)
              .order_by("-fecha_entrada", "-id")
              .first())
    if previo:
        return previo.numero_mesa

    web_user_id = request.session.get("web_user_id")
    if web_user_id:
        usuario = Usuario.objects.filter(id=web_user_id).only("mesa").first()
        if usuario:
            mesa_usuario = _to_positive_int(usuario.mesa)
            if mesa_usuario:
                return mesa_usuario

    return 1

@web_admin_required
def inicio(request):
    disponibilidades = Disponibilidad.objects.all()
    mesas_disponibles = sorted(
        {
            str(m).strip()
            for m in Mesa.objects.values_list("nombre", flat=True)
            if str(m).strip()
        },
        key=lambda x: (0, int(x)) if x.isdigit() else (1, x.lower()),
    )
    return render(request, 'disponibilidad.html', {
        'disponibilidades': disponibilidades,
        'mesas_disponibles': mesas_disponibles,
    })


@web_admin_required
def eliminar_disponibilidad(request, id):
    try:
        Disponibilidad.objects.get(id=id).delete()
        messages.success(request, "Disponibilidad eliminada correctamente")
    except Disponibilidad.DoesNotExist:
        messages.error(request, "La disponibilidad no existe.")
    return redirect('dispo')

@web_admin_required
@require_POST
def procesar_edicion_disponibilidad(request):
    if request.method == "POST":
        try:
            _id = (request.POST.get("id") or "").strip()
            stock_raw = (request.POST.get("stock") or "").strip()
            if stock_raw == "":
                raise ValueError("El stock es obligatorio.")

            try:
                stock = int(stock_raw)
            except (TypeError, ValueError):
                raise ValueError("El stock debe ser un numero entero.")

            if stock < 0:
                raise ValueError("El stock no puede ser menor a 0.")

            fecha_entrada_raw = (request.POST.get("fecha_entrada") or "").strip()
            fecha_entrada_date = None
            if fecha_entrada_raw:
                try:
                    f = fecha_entrada_raw
                    # support "Z" timezone suffix from ISO strings
                    if f.endswith("Z"):
                        f = f[:-1] + "+00:00"
                    fecha_entrada_date = datetime.fromisoformat(f).date()
                except Exception:
                    fecha_entrada_date = None

            is_new = False
            if _id:
                d = Disponibilidad.objects.get(id=_id)  # edita
                d.stock = stock
            else:
                variedad = (request.POST.get("variedad") or "").strip()
                medida = (request.POST.get("medida") or "").strip()
                mesa_raw = request.POST.get("numero_mesa")

                if not variedad or not medida:
                    raise ValueError("Faltan variedad y medida para crear el registro.")

                mesa = _resolver_mesa_para_creacion(request, variedad, medida, mesa_raw)

                # intenta reutilizar un registro existente para evitar crear duplicados
                qs = Disponibilidad.objects.filter(
                    numero_mesa=mesa,
                    variedad=variedad,
                    medida=medida,
                )
                if fecha_entrada_date:
                    qs = qs.filter(fecha_entrada__date=fecha_entrada_date)
                else:
                    qs = qs.filter(fecha_entrada__date=timezone.localdate())

                existing = qs.order_by("-fecha_entrada", "-id").first()
                if existing:
                    d = existing
                    d.stock = stock
                else:
                    is_new = True
                    d = Disponibilidad(
                        numero_mesa=mesa,
                        variedad=variedad,
                        medida=medida,
                        stock=stock,
                        fecha_entrada=timezone.now(),
                    )

            # Si se llega a 0, marcamos la salida (o limpiamos si vuelve a tener stock)
            if stock == 0:
                d.fecha_salida = timezone.now()
            else:
                d.fecha_salida = None

            d.save()

            msg = "Disponibilidad creada correctamente" if is_new else "Disponibilidad actualizada correctamente"

            # ==========================
            #  WEBSOCKET igual que antes
            # ==========================
            async_to_sync(get_channel_layer().group_send)(
                "disponibilidad",
                {
                    "type": "nueva_disponibilidad",
                    "data": DisponibilidadSerializer(d).data
                }
            )

            messages.success(request, msg)

        except Exception as e:
            messages.error(request, f"Error: {e}")

        return redirect('dispo')



# =========================
# API REST – VIEWSET
# =========================

class DisponibilidadViewSet(viewsets.ModelViewSet):
    queryset = Disponibilidad.objects.all().order_by('-fecha_entrada')
    serializer_class = DisponibilidadSerializer

    @action(detail=False, methods=['get'])
    def activos(self, request):
        qs = Disponibilidad.objects.filter(fecha_salida__isnull=True)
        return Response(self.get_serializer(qs, many=True).data)

    @action(detail=False, methods=['get'])
    def por_mesa(self, request):
        mesa = request.query_params.get("mesa")
        if not mesa:
            return Response({"error": "Parámetro mesa requerido"}, status=400)
        qs = Disponibilidad.objects.filter(numero_mesa=mesa)
        return Response(self.get_serializer(qs, many=True).data)


# =========================
# API MANUAL
# =========================
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def api_disponibilidad_list(request):
    print(" user:", request.user, "auth:", request.user.is_authenticated)
    print(" cookies:", request.COOKIES)
    print("session keys:", list(request.session.keys()))

    if request.method == 'GET':
        ordenar = request.query_params.get("ordenar")
        fecha = request.query_params.get("fecha")
        mesa = request.query_params.get("mesa")
        desde = request.query_params.get("desde")
        hasta = request.query_params.get("hasta")
        reciente = request.query_params.get("reciente")

        # Validación de fechas (solo desde 2026-03-06 en adelante)
        MIN_DATE = datetime(2026, 3, 6).date()
        hoy = timezone.localdate()

        def _parse_fecha(fecha_str):
            try:
                return datetime.strptime(fecha_str, "%Y-%m-%d").date()
            except Exception:
                return None

        if fecha:
            fecha_dt = _parse_fecha(fecha)
            if not fecha_dt:
                return Response({"error": "Fecha inválida."}, status=400)
            if fecha_dt < MIN_DATE or fecha_dt > hoy:
                return Response({"error": f"La fecha debe ser entre {MIN_DATE} y {hoy}."}, status=400)

        if desde:
            desde_dt = _parse_fecha(desde)
            if not desde_dt:
                return Response({"error": "Fecha 'desde' inválida."}, status=400)
            if desde_dt < MIN_DATE or desde_dt > hoy:
                return Response({"error": f"La fecha 'desde' debe ser entre {MIN_DATE} y {hoy}."}, status=400)

        if hasta:
            hasta_dt = _parse_fecha(hasta)
            if not hasta_dt:
                return Response({"error": "Fecha 'hasta' inválida."}, status=400)
            if hasta_dt < MIN_DATE or hasta_dt > hoy:
                return Response({"error": f"La fecha 'hasta' debe ser entre {MIN_DATE} y {hoy}."}, status=400)

        qs = Disponibilidad.objects.all()

        if mesa:
            mesa_int = _to_positive_int(mesa)
            if mesa_int is None:
                qs = qs.none()
            else:
                qs = qs.filter(numero_mesa=mesa_int)
        
        if fecha:
            qs = qs.filter(fecha_entrada__date=fecha)

        if desde and hasta:
            qs = qs.filter(fecha_entrada__date__range=[desde, hasta])

        campos = {
            "mesa": "numero_mesa",
            "variedad": "variedad",
            "medida": "medida",
            "fecha": "fecha_entrada"
        }

        if ordenar in campos:
            campo = campos[ordenar]
            if reciente == "true":
                campo = "-" + campo
            qs = qs.order_by(campo)
        
        return Response(DisponibilidadSerializer(qs, many=True).data)

    elif request.method == 'POST':

        data = request.data
        codigo = data.get("qr_id")
        mesa = data.get("numero_mesa")
        variedad = data.get("variedad")
        medida = data.get("medida")

        if not codigo or not mesa or not variedad or not medida:
            return Response({"error": "Datos incompletos"}, status=status.HTTP_400_BAD_REQUEST)


        if QRDisponibilidadUsado.objects.filter(qr_id=codigo).exists():
            return Response(
                {"error": "Este QR ya fue utilizado en Disponibilidad"},
                status=status.HTTP_409_CONFLICT
            )

        # Guardar QR para siempre (protegido ante concurrencia)
        try:
            QRDisponibilidadUsado.objects.create(qr_id=codigo)
        except IntegrityError:
            return Response(
                {"error": "Este QR ya fue utilizado en Disponibilidad"},
                status=status.HTTP_409_CONFLICT
            )

        hoy = timezone.localdate()

        existente = Disponibilidad.objects.filter(
            numero_mesa=mesa,
            variedad=variedad,
            medida=medida,
            fecha_entrada__date=hoy
        ).first()

        if existente:
            existente.stock += 1

            #  CLAVE: si estaba cerrada porque llegó a 0, reabrirla
            existente.fecha_salida = None

            existente.save()


            async_to_sync(get_channel_layer().group_send)(
                "disponibilidad",
                {
                    "type": "nueva_disponibilidad",
                    "data": DisponibilidadSerializer(existente).data
                }
            )
            return Response(DisponibilidadSerializer(existente).data, status=200)

        nuevo = Disponibilidad.objects.create(
            numero_mesa=mesa,
            variedad=variedad,
            medida=medida,
            stock=1,
            fecha_entrada=timezone.now()
        )

        async_to_sync(get_channel_layer().group_send)(
            "disponibilidad",
            {
                "type": "nueva_disponibilidad",
                "data": DisponibilidadSerializer(nuevo).data
            }
        )

        return Response(DisponibilidadSerializer(nuevo).data, status=201)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def api_disponibilidad_detail(request, pk):
    try:
        disponibilidad = Disponibilidad.objects.get(pk=pk)
    except Disponibilidad.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = DisponibilidadSerializer(disponibilidad)
        return Response(serializer.data)

    elif request.method == 'PUT':
        serializer = DisponibilidadSerializer(disponibilidad, data=request.data)
        if serializer.is_valid():
            serializer.save()

            async_to_sync(get_channel_layer().group_send)(
                "disponibilidad",
                {
                    "type": "nueva_disponibilidad",
                    "data": serializer.data
                }
            )

            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        disponibilidad.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_disponibilidad_stats(request):
    return Response({
        "total_registros": Disponibilidad.objects.count(),
        "registros_activos": Disponibilidad.objects.filter(fecha_salida__isnull=True).count(),
        "stock_total": Disponibilidad.objects.aggregate(Sum('stock'))['stock__sum'] or 0,
        "mesas_activas": Disponibilidad.objects.filter(fecha_salida__isnull=True)
                            .values('numero_mesa')
                            .distinct()
                            .count()
    })
#API PARA LA DISPONIBILIDAD QUE SALE
from .models import Disponibilidad, QRDisponibilidadSalidaUsado

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_disponibilidad_salida(request):
    data = request.data
    codigo = (data.get("qr_id") or "").strip()
    mesa = data.get("numero_mesa")
    variedad = (data.get("variedad") or "").strip()
    medida = (data.get("medida") or "").strip()

    mesa_int = _to_positive_int(mesa)

    if not codigo or mesa_int is None or not variedad or not medida:
        return Response(
            {"error": "Datos incompletos: qr_id, numero_mesa, variedad y medida son obligatorios"},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Normalizar para evitar problemas de mayúsculas/minúsculas
    variedad_norm = variedad.strip()
    medida_norm = medida.strip()

    base_qs = Disponibilidad.objects.filter(
        numero_mesa=mesa_int,
        variedad__iexact=variedad_norm,
        medida__iexact=medida_norm,
    )

    if not base_qs.exists():
        return Response(
            {"error": "No existe registro de disponibilidad para esa combinación de mesa/variedad/medida."},
            status=status.HTTP_404_NOT_FOUND
        )

    #  Si ya se restó este QR una vez, NO permitir otra vez
    if QRDisponibilidadSalidaUsado.objects.filter(qr_id=codigo).exists():
        return Response(
            {"error": "Este QR ya fue utilizado en SALIDA (ya se restó una vez)"},
            status=status.HTTP_409_CONFLICT
        )

    # Si hay registros con stock > 0, usaremos el más antiguo (FIFO).
    # No dependemos de fecha_salida: si hay stock y está cerrada, la reabrimos.
    if not base_qs.filter(stock__gt=0).exists():
        return Response(
            {"error": "No hay stock disponible para esa variedad y medida"},
            status=status.HTTP_409_CONFLICT
        )

    channel_layer = get_channel_layer()

    with transaction.atomic():
        dispo = (Disponibilidad.objects
                 .select_for_update()
                 .filter(
                    numero_mesa=mesa_int,
                    variedad__iexact=variedad_norm,
                    medida__iexact=medida_norm,
                    stock__gt=0
                )
                 .order_by('fecha_entrada', 'id')
                 .first())

        # Si el registro estaba cerrado, lo reabrimos antes de descontar.
        if dispo and dispo.fecha_salida is not None:
            dispo.fecha_salida = None
            dispo.save()

        if not dispo:
            return Response(
                {"error": "No hay stock disponible para esa variedad y medida"},
                status=status.HTTP_409_CONFLICT
            )

        # Registrar QR como ya restado (protegido ante concurrencia)
        try:
            QRDisponibilidadSalidaUsado.objects.create(qr_id=codigo)
        except IntegrityError:
            return Response(
                {"error": "Este QR ya fue utilizado en SALIDA (ya se restó una vez)"},
                status=status.HTTP_409_CONFLICT
            )

        # Restar 1
        dispo.stock -= 1

        # Si llega a 0, marca fecha_salida (opcional)
        if dispo.stock == 0:
            dispo.fecha_salida = timezone.now()

        dispo.save()

    # Notificar por websocket
    async_to_sync(channel_layer.group_send)(
        "disponibilidad",
        {
            "type": "nueva_disponibilidad",
            "data": DisponibilidadSerializer(dispo).data
        }
    )

    return Response(DisponibilidadSerializer(dispo).data, status=status.HTTP_200_OK)
################################
#API VARIEDAD#
################################
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def variedades_api(request):
    if request.method == "GET":
        qs = Variedad.objects.all().order_by("nombre")
        return Response(VariedadSerializer(qs, many=True).data)

    nombre_raw = (request.data.get("nombre") or "").strip()
    if not nombre_raw:
        return Response({"detail": "El nombre es obligatorio."}, status=400)

    #  Normalizar (Explorer)
    nombre = nombre_raw.lower().capitalize()

    #  Mensaje desde el backend
    if Variedad.objects.filter(nombre__iexact=nombre).exists():
        return Response(
            {"detail": "La variedad ya se encuentra agregada."},
            status=409
        )

    nueva = Variedad.objects.create(nombre=nombre)
    return Response(
        {"detail": "Variedad agregada correctamente.", "variedad": VariedadSerializer(nueva).data},
        status=201
    )



from openpyxl import load_workbook
from io import BytesIO

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def variedades_excel_api(request):
    """
    Recibe un archivo Excel con variedades.
    Formatos aceptados:
    - Columna con encabezado: 'variedad'
    - O primera columna sin encabezado
    """
    file = request.FILES.get("file")
    if not file:
        return Response({"detail": "Debes enviar un archivo en 'file'."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        wb = load_workbook(filename=BytesIO(file.read()), data_only=True)
        ws = wb.active

        # leer filas
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({"detail": "El Excel está vacío."}, status=status.HTTP_400_BAD_REQUEST)

        # detectar header
        header = [str(x).strip().lower() if x is not None else "" for x in rows[0]]
        idx = None
        if "variedad" in header:
            idx = header.index("variedad")
            data_rows = rows[1:]
        else:
            idx = 0
            data_rows = rows

        nombres = []
        for r in data_rows:
            if not r or len(r) <= idx:
                continue
            val = r[idx]
            if val is None:
                continue
            nombre = str(val).strip()
            if nombre:
                nombres.append(nombre)

        # quitar duplicados (ignorando mayúsculas)
        unicos = []
        seen = set()
        for n in nombres:
            k = n.lower()
            if k not in seen:
                seen.add(k)
                unicos.append(n)

        creadas = 0
        existentes = 0

        for n in unicos:
            if Variedad.objects.filter(nombre__iexact=n).exists():
                existentes += 1
            else:
                Variedad.objects.create(nombre=n)
                creadas += 1

        return Response({
            "detail": "ok",
            "creadas": creadas,
            "existentes": existentes,
            "total": len(unicos)
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({"detail": f"Error leyendo Excel: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)


from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from rest_framework.response import Response
from openpyxl import load_workbook
from io import BytesIO

class VariedadViewSet(viewsets.ModelViewSet):
    queryset = Variedad.objects.all().order_by("nombre")
    serializer_class = VariedadSerializer
    permission_classes = [IsAuthenticated]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()

        # Permite borrar la variedad si no existe stock activo (> 0).
        # Si todos los registros historicos estan en 0, se puede eliminar.
        tiene_stock_activo = Disponibilidad.objects.filter(
            variedad__iexact=instance.nombre,
            stock__gt=0
        ).exists()
        if tiene_stock_activo:
            return Response(
                {"detail": "No puedes borrar esta variedad porque tiene stock mayor a 0."},
                status=status.HTTP_409_CONFLICT
            )
        return super().destroy(request, *args, **kwargs)

    #  AQUÍ el excel SIN choque con /<pk>/
    @action(detail=False, methods=["post"], url_path="excel")
    def excel(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "Debes enviar un archivo en 'file'."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = load_workbook(filename=BytesIO(file.read()), data_only=True)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return Response({"detail": "El Excel está vacío."}, status=status.HTTP_400_BAD_REQUEST)

            header = [str(x).strip().lower() if x is not None else "" for x in rows[0]]
            if "variedad" in header:
                idx = header.index("variedad")
                data_rows = rows[1:]
            else:
                idx = 0
                data_rows = rows

            nombres = []
            for r in data_rows:
                if not r or len(r) <= idx:
                    continue
                val = r[idx]
                if val is None:
                    continue
                nombre = str(val).strip()
                if nombre:
                    nombres.append(nombre)

            # quitar duplicados por lower
            seen = set()
            unicos = []
            for n in nombres:
                k = n.lower()
                if k not in seen:
                    seen.add(k)
                    unicos.append(n)

            creadas = 0
            existentes = 0

            for n in unicos:
                if Variedad.objects.filter(nombre__iexact=n).exists():
                    existentes += 1
                else:
                    Variedad.objects.create(nombre=n)
                    creadas += 1

            return Response({
                "detail": "ok",
                "creadas": creadas,
                "existentes": existentes,
                "total": len(unicos)
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"detail": f"Error leyendo Excel: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated]) 
def listar_variedades_api(request):
    variedades = Variedad.objects.all().order_by('nombre')
    data = [{'id': v.id, 'nombre': v.nombre} for v in variedades]
    return Response({'success': True, 'data': data, 'count': len(data)})
